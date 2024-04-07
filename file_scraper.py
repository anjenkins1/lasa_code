from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import os

def get_data(filename, excel_file_path):
    # Read the HTML file
    with open(filename, 'r') as file:
        html_content = file.read()

    # Parse the HTML content
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find all div elements with class "address-container"
    address_containers = soup.find_all('div', class_='address-container')

    # Create a new Excel workbook and select the active worksheet
    try:
        wb = load_workbook(excel_file_path)
    except FileNotFoundError:
        wb = Workbook()
        
    # Select the active worksheet
    ws = wb.active

    # Set headers
    if ws.max_row == 0:
        ws.append(["Name", "Address"])

    # Iterate over the address containers and extract data
    for container in address_containers:
        # Extract the name (h4 tag)
        name = container.find('h4').text.strip()
        
        # Extract the address (address tag)
        address = container.find('address').text.strip().replace('\n', ', ')
        
        # Write data to Excel
        ws.append([name, address])

    # Save the Excel file
    wb.save(excel_file_path)


# Directory containing HTML files
directory = 'pages_to_scrape_ISH-WISH'

# Iterate over each file in the directory
for filename in os.listdir(directory):
    get_data(os.path.join(directory, filename), "/home/jenk2319/lasa_code/addresses.xlsx")